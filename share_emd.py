"""闲鱼接单文案生成 → Telegram 推送（纯文字版）"""

import os
import random
import requests
from datetime import datetime

import openpyxl
from dotenv import load_dotenv

load_dotenv()


# =============================
# 配置（优先读环境变量）
# =============================

XLSX_PATH = os.getenv("SHARE_XLSX_PATH", "data/share_modules.xlsx")

TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "").strip()
TELEGRAM_CHAT_ID = os.getenv("TELEGRAM_CHAT_ID", "").strip()

QWEN_API_KEY = os.getenv("QWEN_API_KEY", os.getenv("DASHSCOPE_API_KEY", "")).strip()
QWEN_BASE_URL = os.getenv("QWEN_BASE_URL", "https://dashscope.aliyuncs.com/compatible-mode/v1").rstrip("/")
QWEN_MODEL = os.getenv("QWEN_MODEL", "qwen3-max-2026-01-23")


# =============================
# Excel 读取（用 openpyxl，不依赖 pandas）
# =============================

def load_sheets(path: str) -> dict:
    """读取 xlsx，返回 {sheet_name: [rows]}"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h).strip().lower() if h else f"col{i}" for i, h in enumerate(rows[0])]
        data = []
        for row in rows[1:]:
            d = {headers[i]: (str(v).strip() if v else "") for i, v in enumerate(row) if i < len(headers)}
            if any(d.values()):
                data.append(d)
        result[name.lower()] = data
    wb.close()
    return result


def pick_one(rows: list[dict], col: str) -> str:
    return random.choice(rows)[col]


def pick_many(rows: list[dict], col: str, k: int) -> list[str]:
    k = min(k, len(rows))
    return [r[col] for r in random.sample(rows, k) if r.get(col)]


# =============================
# 文案组装
# =============================

def build_raw_text(sheets: dict) -> str:
    modules = sheets.get("modules", [])

    if not modules:
        raise ValueError("modules sheet 为空")

    def pick_type(t: str) -> str:
        sub = [r for r in modules if r.get("type", "").lower() == t]
        if not sub:
            raise ValueError(f"modules 里找不到 type={t}")
        return pick_one(sub, "text")

    hook = pick_type("hook")
    ability = pick_type("ability")
    service = pick_type("service")
    deliver1 = pick_type("deliver")
    deliver2 = pick_type("deliver")
    trust = pick_type("trust")
    cta = pick_type("cta")

    # 防止 deliver 重复
    deliver_pool = [r for r in modules if r.get("type", "").lower() == "deliver"]
    if deliver2 == deliver1 and len(deliver_pool) > 1:
        for _ in range(5):
            d = pick_type("deliver")
            if d != deliver1:
                deliver2 = d
                break

    # 关键词服务范围（替代具体项目列举）
    keyword_services = [
        "单片机设计代做", "stm32单片机设计", "stm32单片机代做",
        "esp32单片机", "51单片机设计", "单片机程序开发",
        "单片机定制开发", "单片机实物设计", "arduino开发",
        "电路设计", "pcb设计", "电路仿真设计", "嘉立创打板",
        "单片机项目开发", "单片机定制", "单片机程序设计",
        "keil程序开发", "zigbee开发", "lora开发",
        "硬件实物", "软件开发定制", "app开发", "云平台对接",
    ]
    kw_sample = random.sample(keyword_services, min(8, len(keyword_services)))

    raw = (
        f"{hook}\n\n"
        f"{ability}\n"
        f"{service}\n\n"
    )
    raw += f"可接范围：{' / '.join(kw_sample)}\n\n"
    raw += (
        f"交付：\n"
        f"- {deliver1}\n"
        f"- {deliver2}\n\n"
        f"{trust}\n\n"
        f"{cta}"
    )
    return raw


# =============================
# Qwen 润色
# =============================

def qwen_polish(text: str) -> str:
    if not QWEN_API_KEY:
        return text

    url = f"{QWEN_BASE_URL}/chat/completions"
    headers = {
        "Authorization": f"Bearer {QWEN_API_KEY}",
        "Content-Type": "application/json",
    }

    user_prompt = (
        '你是闲鱼SEO文案助手。目标：让买家搜索关键词时能搜到这条文案。\n\n'
        '核心搜索关键词（必须尽量多地自然嵌入，每条文案至少命中5个不同的关键词）：\n'
        '单片机设计代做、单片机代做、单片机设计、单片机定制、单片机项目代做、\n'
        'stm32单片机设计、stm32单片机代做、单片机程序开发、单片机程序设计、\n'
        '单片机定制开发、单片机实物设计、esp32单片机、51单片机、arduino开发、\n'
        '电路设计、pcb设计、电路仿真设计、嘉立创打板、单片机项目开发\n\n'
        '规则（必须遵守）：\n'
        '1) 关键词优先：文案围绕上面的搜索关键词组织，用关键词本身描述服务能力，'
        '禁止出现具体项目名称（禁止写"xx系统""xx机器""智能xx"等具体项目描述，会稀释关键词权重）。\n'
        '2) 不要自称"资深工程师""专业团队""多年经验"等包装人设，用朴实口吻，像普通卖家发帖。\n'
        '3) 服务范围用关键词表达，例如：接单片机设计代做、stm32单片机程序开发、'
        'esp32单片机代做、51单片机设计、电路设计、pcb设计、嘉立创打板、硬件实物等。\n'
        '4) 禁止出现：课程设计、毕业设计、大学生、学术代做等字眼。\n'
        '5) 禁止夸张词（轻松搞定/从0到1/完美/秒出/全网最低）和强营销口号（限时/速来/赶紧下单）。\n'
        '6) 保留一句自然的私信引导（如"有需要可以私信聊"）。\n'
        '7) 不要使用小标题（如【交付内容】），用自然段。\n'
        '8) 字数控制在 220~320 字。\n\n'
        '把下面文本润色后输出（只输出润色后的最终文案）：\n\n'
        f'<<<\n{text}\n>>>'
    )

    payload = {
        "model": QWEN_MODEL,
        "messages": [
            {"role": "system", "content": "你是闲鱼SEO获客文案助手，擅长把搜索关键词自然嵌入文案，风格朴实像真人发帖。"},
            {"role": "user", "content": user_prompt}
        ],
        "temperature": 0.7
    }

    r = requests.post(url, headers=headers, json=payload, timeout=60)
    r.raise_for_status()
    return r.json()["choices"][0]["message"]["content"].strip()


# =============================
# Telegram 发送
# =============================

def telegram_send_text(msg: str):
    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID:
        print(f"[Telegram 未配置] {msg[:80]}...")
        return

    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    payload = {"chat_id": TELEGRAM_CHAT_ID, "text": msg}
    r = requests.post(url, json=payload, timeout=30)
    r.raise_for_status()


# =============================
# 主流程：生成 N 条 + 汇报
# =============================

def run_batch(count: int = 4):
    """生成 count 条文案，带编号发送，最后发日期汇报"""
    sheets = load_sheets(XLSX_PATH)

    success = 0
    for i in range(1, count + 1):
        try:
            raw = build_raw_text(sheets)
            final = qwen_polish(raw) if QWEN_API_KEY else raw

            # 加编号前缀
            numbered = f"【{i}/{count}】\n\n{final}"
            telegram_send_text(numbered)
            success += 1
            print(f"[{i}/{count}] 已发送")
        except Exception as e:
            print(f"[{i}/{count}] 失败: {e}")

    # 发送日期汇报
    today = datetime.now().strftime("%Y-%m-%d %H:%M")
    report = f"--- 今日文案推送完毕 ---\n日期: {today}\n成功: {success}/{count} 条"
    telegram_send_text(report)
    print(report)


if __name__ == "__main__":
    import sys
    n = int(sys.argv[1]) if len(sys.argv) > 1 else 4
    run_batch(n)
